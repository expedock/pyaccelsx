import sys
import openpyxl

# Default to 1 sheet with 1000 rows x 50 cols
row_max = 4000
col_max = 50
sheets = 1

for i, arg in enumerate(sys.argv):
    if i == 1:
        row_max = int(sys.argv[1])
    elif i == 2:
        col_max = int(sys.argv[2])
    elif i == 3:
        sheets = int(sys.argv[3])

workbook = openpyxl.Workbook()

for r in range(sheets):
    worksheet = workbook.create_sheet()
    for row in range(1, row_max + 1):
        for col in range(1, col_max + 1):
            if col % 2:
                worksheet.cell(row, col).value = "Foo"
            else:
                worksheet.cell(row, col).value = 12345

workbook.save("openpyxl_perf_test.xlsx")